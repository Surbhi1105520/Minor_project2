import pytest
import os
import datetime as dt
import allure
from pytest_html import extras
from selenium.common.exceptions import TimeoutException, WebDriverException
from openpyxl import load_workbook
from selenium import webdriver
from selenium.webdriver.chrome.service import Service
from selenium.webdriver.common.by import By
from selenium.webdriver.support.ui import WebDriverWait
from selenium.webdriver.support import expected_conditions as EC

DEFAULT_EXCEL = "D:\GUVI\Project\sample1.xlsx"
#DEFAULT_EXCEL = "D:\GUVI\TASK\T15\sample1.xlsx"
DEFAULT_SHEET = "Sheet1"
LOGIN_URL = "https://opensource-demo.orangehrmlive.com/web/index.php/auth/login"



# ---------- CLI options (defaults to sample1.xlsx / Sheet1) ----------
def pytest_addoption(parser):
    parser.addoption("--excel", action="store", default=DEFAULT_EXCEL,
                     help="Path to Excel workbook (default: D:\GUVI\Project\sample1.xlsx)")
    parser.addoption("--sheet", action="store", default=DEFAULT_SHEET,
                     help="Worksheet name (default: Sheet1)")


@pytest.fixture(scope="session")
def excel_path(request):
    path = request.config.getoption("--excel")
    if not os.path.exists(path):
        raise FileNotFoundError(f"Excel file not found: {path}")
    return path


@pytest.fixture(scope="session")
def sheet_name(request):
    return request.config.getoption("--sheet")


@pytest.fixture(scope="session")
def excel_book(excel_path):
    wb = load_workbook(excel_path)
    yield wb
    wb.save(excel_path)  # final safety save


def _read_rows(wb, sheet_name):
    """Read rows into dicts using header names """
    if sheet_name not in wb.sheetnames:
        raise ValueError(f"Sheet '{sheet_name}' not found.")
    ws = wb[sheet_name]

    headers = [c.value for c in ws[1]]
    header_map = {h: i + 1 for i, h in enumerate(headers) if h}

    required = {"S. No.", "TestID", "NameofTester","Date","TestParameter", "Username", "Password","TestResult"}
    if not required.issubset(header_map.keys()):
        missing = required - set(header_map.keys())
        raise ValueError(f"Missing required headers: {missing}")

    cases = [] #not considering first row of labels
    for row in ws.iter_rows(min_row=2, values_only=False):
        testid_cell = row[header_map["TestID"] - 1]
        if testid_cell.value is None:
            continue

        def val(col):
            return row[header_map[col] - 1].value

        cases.append({
            "S. No.": val("S. No."),
            "TestID": val("TestID"),
            "NameofTester": val("NameofTester"),
            "Date": val("Date"),
            "TestParameter": val("TestParameter"),
            "Username": val("Username"),
            "Password": val("Password"),
            "TestResult": val("TestResult"),
            "__rownum": testid_cell.row,
            "__result_col": header_map["TestResult"],
        })
    return cases


def _write_result(wb, sheet_name, rownum, colnum, value):
    ws = wb[sheet_name]
    ws.cell(row=rownum, column=colnum, value=value)


def pytest_generate_tests(metafunc):
    """Parametrize with 'case' from Excel rows."""
    if "case" in metafunc.fixturenames:
        excel = metafunc.config.getoption("--excel")
        sheet = metafunc.config.getoption("--sheet")
        wb = load_workbook(excel)
        cases = _read_rows(wb, sheet)
        ids = [f"Row{c['TestID']}:{c['Username']}" for c in cases]
        metafunc.parametrize("case", cases, ids=ids)


@pytest.fixture
def driver():
    driver = webdriver.Chrome(service=Service())
    driver.maximize_window()
    driver.get(LOGIN_URL)
    yield driver
    driver.quit()


@pytest.fixture
def wait(driver):
    return WebDriverWait(driver, 10)


@pytest.fixture
def result_writer(excel_book, excel_path, sheet_name):
    def _write(case, value: str):
        _write_result(excel_book, sheet_name, case["__rownum"], case["__result_col"], value)
        excel_book.save(excel_path)  # save after each write
    return _write

def _attach_allure(driver, name_prefix="on_fail"):
    try:
        allure.attach(
            driver.get_screenshot_as_png(),
            name=f"{name_prefix}_screenshot",
            attachment_type=allure.attachment_type.PNG,
        )
    except Exception:
        pass
    try:
        allure.attach(
            driver.page_source,
            name=f"{name_prefix}_page_source",
            attachment_type=allure.attachment_type.HTML,
        )
    except Exception:
        pass
    try:
        allure.attach(
            driver.current_url,
            name=f"{name_prefix}_url",
            attachment_type=allure.attachment_type.TEXT,
        )
    except Exception:
        pass

@pytest.hookimpl(tryfirst=True, hookwrapper=True)
def pytest_runtest_makereport(item, call):
    """Add HTML + Allure attachments for failures, and put test docstring as description in HTML."""
    outcome = yield
    report = outcome.get_result()

    # show function docstring as description in pytest-html
    report.description = (item.function.__doc__ or "").strip()

    if report.when != "call":
        return

    if report.failed and "driver" in item.fixturenames:
        driver = item.funcargs["driver"]

        # ---- Allure attachments ----
        _attach_allure(driver, name_prefix="fail")

        # ---- pytest-html attachments ----
        extra = getattr(report, "extra", [])
        try:
            # screenshot as base64
            extra.append(extras.image(driver.get_screenshot_as_base64(), "Screenshot on failure"))
        except WebDriverException:
            pass
        try:
            extra.append(extras.html(f"<div><b>URL:</b> {driver.current_url}</div>"))
        except Exception:
            pass
        report.extra = extra

def pytest_html_report_title(report):
    report.title = "Mini_Project2 — Test Report"

def pytest_configure(config):
    """Add metadata to HTML report and create Allure environment.properties if --alluredir is set."""
    # pytest-metadata plugin (already in your env) is picked up by pytest-html
    md = getattr(config, "_metadata", {})
    md["Project"] = "Mini_Project2"
    md["Base URL"] = "https://opensource-demo.orangehrmlive.com"
    md["Run At"] = dt.datetime.now().strftime("%Y-%m-%d %H:%M:%S")

    # Write Allure environment.properties to the chosen --alluredir
    adir = config.getoption("--alluredir")
    if adir:
        os.makedirs(adir, exist_ok=True)
        env_file = os.path.join(adir, "environment.properties")
        with open(env_file, "w", encoding="utf-8") as f:
            f.write("Project=Mini_Project2\n")
            f.write("BaseURL=https://opensource-demo.orangehrmlive.com\n")
            f.write("Browser=Chrome\n")
            f.write(f"RunAt={dt.datetime.now().isoformat(timespec='seconds')}\n")
